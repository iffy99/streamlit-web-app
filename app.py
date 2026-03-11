import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
import re
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

# --- Page Config ---
st.set_page_config(page_title="AWW Screening Records", layout="wide")

# --- 1. MEMORY LOGIC ---
if 'all_selected_awcs' not in st.session_state:
    st.session_state.all_selected_awcs = []
if 'sel_icds' not in st.session_state:
    st.session_state.sel_icds = []

# --- Custom CSS ---
st.markdown("""
    <style>
    @keyframes blinker { 50% { border-color: #ff4b4b; box-shadow: 0 0 10px #ff4b4b; } }
    .blink-box div[data-baseweb="select"] { border: 2px solid #ff4b4b !important; animation: blinker 1s linear infinite; border-radius: 4px; }
    .selection-summary { background-color: #f0f2f6; padding: 15px; border-radius: 10px; border-left: 5px solid #0068c9; margin-bottom: 20px; }
    .summary-tag { display: inline-block; background-color: #0068c9; color: white; padding: 2px 8px; border-radius: 4px; margin: 2px; font-size: 11px; }
    [data-testid="stMetricValue"] { font-size: 24px; color: #0068c9; }
    </style>
""", unsafe_allow_html=True)

st.title("🚀 AWW Screening Records")

uploaded_file = st.file_uploader("📂 Upload Master Data File", type=['xlsx', 'csv'])

if uploaded_file is not None:
    try:
        removed_count = 0
        with st.status("⏳ Processing Source Data...", expanded=True) as status:
            df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
            
            # --- Cleaning & Mappings ---
            if 'scr_outcome_lbl' in df.columns:
                df['scr_outcome_lbl'] = df['scr_outcome_lbl'].astype(str).str.strip()
                removed_count = len(df[df['scr_outcome_lbl'] == "Outside Age Range"])
                df = df[df['scr_outcome_lbl'] != "Outside Age Range"].copy()
            
            type_map = {'1': "1-Actual(First Screening)", '2': "2-Duplicate Screening", '9': "9-Actual (Repeat Screening)", 'Pending': "Pending Duplicate Check", '0': "0-Test"}
            relation_map = {'1': "Father", '2': "Mother", '3': "Grandparent", '4': "any other family member", '5': "Neighbor", '99': "Other"}

            if 'scr_type' in df.columns:
                df['scr_type'] = df['scr_type'].astype(str).str.strip().replace(['nan','','None'],'Pending')
                df['scr_type'] = df['scr_type'].apply(lambda x: str(int(float(x))) if x.replace('.0','').isdigit() else x)
                df['scr_type'] = df['scr_type'].map(type_map).fillna(df['scr_type'])

            if 'scr_outcome_lbl' in df.columns:
                df.loc[df['scr_outcome_lbl'] == "Screening not completed", 'scr_type'] = "Screening not completed"

            for col in ['coninfo_childname', 'coninfo_respname']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.title().str.strip()

            if 'coninfo_resprelation' in df.columns:
                df['coninfo_resprelation'] = df['coninfo_resprelation'].astype(str).str.replace('.0','', regex=False).map(relation_map).fillna(df['coninfo_resprelation'])

            # BACK-END DATE CONVERSION (Added extra date columns for logic)
            date_cols = ['scr_enddate', 'scr_startdate', 'coninfo_serverdt', 'coninfo_childdob']
            for col in date_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')

            status.update(label=f"✅ Data Ready (Removed {removed_count} Outside Age Range records)", state="complete", expanded=False)

        # --- Filters Logic ---
        st.subheader("🔍 Configuration Filters")
        
        if st.button("🧹 Clear All Selections"):
            st.session_state.all_selected_awcs = []
            st.session_state.sel_icds = []
            st.rerun()

        c1, c2, c3, c4 = st.columns([2, 2, 1.5, 1.5])
        
        with c1:
            icds_options = sorted([x for x in df['coninfo_icds'].dropna().unique() if str(x).strip() != ''])
            st.session_state.sel_icds = st.multiselect("1. Select ICDS Name", options=icds_options, default=st.session_state.sel_icds)
        
        with c2:
            if st.session_state.sel_icds:
                current_icds_awcs = sorted(df[df['coninfo_icds'].isin(st.session_state.sel_icds)]['awc_id'].dropna().unique().tolist())
                valid_selections = [awc for awc in st.session_state.all_selected_awcs if awc in current_icds_awcs]
                temp_sel = st.multiselect("2. Select AWC", options=current_icds_awcs, default=valid_selections)
                st.session_state.all_selected_awcs = temp_sel
            else:
                st.info("💡 Please select ICDS first")
                st.session_state.all_selected_awcs = []

        with c3: start_d = st.date_input("3. Start Date", value=date(2024, 9, 1))
        with c4: end_d = st.date_input("4. End Date", value=date.today())

        if st.session_state.sel_icds or st.session_state.all_selected_awcs:
            st.markdown('<div class="selection-summary">', unsafe_allow_html=True)
            cols_sum = st.columns(2)
            with cols_sum[0]:
                st.markdown(f"**📍 Selected ICDS ({len(st.session_state.sel_icds)}):**")
                st.markdown(" ".join([f'<span class="summary-tag">{i}</span>' for i in st.session_state.sel_icds]), unsafe_allow_html=True)
            with cols_sum[1]:
                st.markdown(f"**🏠 Selected AWCs ({len(st.session_state.all_selected_awcs)}):**")
                st.markdown(" ".join([f'<span class="summary-tag">{a}</span>' for a in st.session_state.all_selected_awcs]), unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        st.write("---")
        all_opts = sorted(df['scr_type'].unique().tolist())
        is_pending = "Pending Duplicate Check" in df[df['awc_id'].isin(st.session_state.all_selected_awcs)]['scr_type'].values if st.session_state.all_selected_awcs else False
        
        st.markdown(f'<div class="{"blink-box" if is_pending else ""}">', unsafe_allow_html=True)
        sel_types = st.multiselect("5. Select Screening Status", options=all_opts, default=all_opts)
        st.markdown('</div>', unsafe_allow_html=True)

        # --- REPORT PREVIEW ---
        if st.session_state.all_selected_awcs:
            report_df = df[df['awc_id'].isin(st.session_state.all_selected_awcs)].copy()
            
            # --- UPDATED DATE FILTER LOGIC (Hidden Priority) ---
            # Hum ek temp target_date column bana rahe hain filter ke liye
            target_date = report_df['scr_enddate'].fillna(report_df['scr_startdate']).fillna(report_df['coninfo_startdate'])
            
            # Ab filter isi target_date par lagega
            report_df = report_df[(target_date.dt.date >= start_d) & (target_date.dt.date <= end_d)]
            
            if sel_types:
                report_df = report_df[report_df['scr_type'].isin(sel_types)]
            else:
                report_df = pd.DataFrame(columns=df.columns)

            if not report_df.empty:
                st.subheader("📊 Report Preview & Quick Counts")
                m_row1 = st.columns(6) 
                m_row1[0].metric("Total Records", len(report_df))
                m_row1[1].metric("Actual Records", len(report_df[report_df['scr_type'].str.contains("Actual", na=False)]))
                m_row1[2].metric("Duplicate Records", len(report_df[report_df['scr_type'] == "2-Duplicate Screening"]))
                m_row1[3].metric("Pending Duplicate Check", len(report_df[report_df['scr_type'] == "Pending Duplicate Check"]))
                m_row1[4].metric("Incomplete Screening", len(report_df[report_df['scr_type'] == "Screening not completed"]))
                m_row1[5].metric("Test", len(report_df[report_df['scr_type'].str.contains("Test", na=False)]))

                header_mapping = {
                    'record_id': 'record_id', 'coninfo_icds': 'ICDS', 'coninfo_sector': 'Sector',
                    'awc_id': 'AWC', 'coninfo_childname': 'ChildName',
                    'coninfo_childdob': 'ChildDob', 'coninfo_respname': 'ResponderName',
                    'coninfo_resprelation': 'ResponderRelation', 'coninfo_phone1': 'Phone 1'
                }
                cols = [c for c in header_mapping.keys() if c in report_df.columns]
                
                preview_display = report_df[cols].copy()
                if 'coninfo_childdob' in preview_display.columns:
                    preview_display['coninfo_childdob'] = preview_display['coninfo_childdob'].dt.strftime('%d-%m-%Y')
                
                st.dataframe(preview_display.rename(columns=header_mapping), use_container_width=True, hide_index=True)

                if st.button("📊 Download Categorized Excel Report", use_container_width=True, type="primary"):
                    awc_prefix = "_".join(st.session_state.all_selected_awcs)
                    is_only_actual = all("Actual" in t for t in sel_types) if sel_types else False
                    suffix = "-Actual Records" if is_only_actual else "-Records"
                    final_filename = f"{awc_prefix}{suffix}.xlsx"

                    buffer = io.BytesIO()
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    excel_export_df = report_df.copy()
                    if 'coninfo_childdob' in excel_export_df.columns:
                        excel_export_df['coninfo_childdob'] = excel_export_df['coninfo_childdob'].dt.strftime('%d-%m-%Y')

                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        for awc in excel_export_df['awc_id'].unique():
                            awc_data = excel_export_df[excel_export_df['awc_id'] == awc]
                            ws_name = re.sub(r'[\\/*?:\[\]]', '', str(awc))[:31]
                            curr_ws = writer.book.create_sheet(ws_name)
                            
                            def write_section(df_sec, title, r_pos):
                                if not df_sec.empty:
                                    curr_ws.merge_cells(start_row=r_pos+1, start_column=1, end_row=r_pos+1, end_column=len(cols))
                                    cell = curr_ws.cell(row=r_pos+1, column=1, value=title)
                                    cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center')
                                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                                    cell.border = thin_border
                                    
                                    final_sec = df_sec[cols].rename(columns=header_mapping)
                                    final_sec.to_excel(writer, sheet_name=ws_name, startrow=r_pos+1, index=False)
                                    
                                    for r in range(r_pos + 2, r_pos + len(final_sec) + 3):
                                        for c in range(1, len(cols) + 1):
                                            curr_ws.cell(row=r, column=c).border = thin_border
                                            
                                    return r_pos + len(df_sec) + 3
                                return r_pos

                            cursor = 0
                            cursor = write_section(awc_data[awc_data['scr_type'].str.contains("Actual", na=False)], "ACTUAL RECORDS", cursor)
                            cursor = write_section(awc_data[awc_data['scr_type'] == "2-Duplicate Screening"], "DUPLICATE RECORDS", cursor)
                            cursor = write_section(awc_data[awc_data['scr_type'] == "Screening not completed"], "INCOMPLETE SCREENING RECORDS", cursor)
                            cursor = write_section(awc_data[awc_data['scr_type'] == "Pending Duplicate Check"], "PENDING DUPLICATE CHECK", cursor)
                            cursor = write_section(awc_data[awc_data['scr_type'].str.contains("Test", na=False)], "TEST RECORDS", cursor)

                        # Hata rahe hain default empty sheet ko (agar openpyxl banata hai toh)
                        if 'Sheet' in writer.book.sheetnames:
                            writer.book.remove(writer.book['Sheet'])

                    st.download_button(label="📥 Click to Download Report", data=buffer.getvalue(), file_name=final_filename, use_container_width=True)
            else:
                st.warning("⚠️ No records found.")
    except Exception as e:
        st.error(f"Error: {e}")